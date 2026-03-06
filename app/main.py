"""
FastAPI application — routes for upload, query, and stats.
"""
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import csv
import io

from app.agent import run_agent
from app.datasets import store_dataset, get_dataset, list_datasets, DATASETS
from app.config import settings

# Optional: keep sheets import for legacy endpoint
try:
    from app.sheets import get_sheet_data, sheet_to_objects
    _HAS_SHEETS = True
except Exception:
    _HAS_SHEETS = False

app = FastAPI(title="SheetSearch AI — Structured Query Engine")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Request models ──

class AgentAskRequest(BaseModel):
    question: str
    dataset_id: str | None = None
    limit: int | None = 10


# ── Google Sheets fallback ──

_DEFAULT_DATASET_ID: str | None = None

def _get_or_create_sheets_dataset() -> str:
    """Load Google Sheets data into a dataset if not already loaded. Returns dataset_id."""
    global _DEFAULT_DATASET_ID
    if _DEFAULT_DATASET_ID and _DEFAULT_DATASET_ID in DATASETS:
        return _DEFAULT_DATASET_ID
    if not _HAS_SHEETS:
        raise RuntimeError("No dataset_id provided and Google Sheets is not configured.")
    raw = get_sheet_data(settings.SHEET_ID, "Products!A1:Z1000")
    rows = sheet_to_objects(raw)
    if not rows:
        raise RuntimeError("Google Sheets returned no data.")
    _DEFAULT_DATASET_ID = store_dataset(rows, "Google Sheets - Products")
    return _DEFAULT_DATASET_ID


# ── Upload endpoint ──

def _parse_csv(content: bytes, filename: str) -> list[dict]:
    """Parse CSV bytes into a list of row dicts."""
    text = content.decode("utf-8-sig")  # handles BOM
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    return rows


def _parse_xlsx(content: bytes, filename: str) -> list[dict]:
    """Parse XLSX bytes into a list of row dicts. Requires openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=400, detail="XLSX support requires openpyxl. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter, []))]
    if not headers:
        raise HTTPException(status_code=400, detail="XLSX file has no headers.")

    rows = []
    for row_values in rows_iter:
        row = {}
        for i, h in enumerate(headers):
            val = row_values[i] if i < len(row_values) else None
            row[h] = str(val).strip() if val is not None else ""
        rows.append(row)

    wb.close()
    return rows


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    Upload a CSV or XLSX file. Returns dataset_id, columns, and inferred types.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided.")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail=f"Unsupported file type '.{ext}'. Use CSV or XLSX.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File is empty.")

    try:
        if ext == "csv":
            rows = _parse_csv(content, file.filename)
        else:
            rows = _parse_xlsx(content, file.filename)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="File contains no data rows.")

    dataset_id = store_dataset(rows, file.filename)
    ds = get_dataset(dataset_id)

    return {
        "ok": True,
        "dataset_id": dataset_id,
        "filename": file.filename,
        "columns": ds["columns"],
        "types": ds["types"],
        "row_count": ds["row_count"],
    }


# ── Agent query endpoints ──

@app.post("/agent/ask")
def agent_ask(req: AgentAskRequest):
    """Execute a natural language query against a stored dataset."""
    dataset_id = req.dataset_id
    if not dataset_id:
        return {
            "ok": False,
            "answer_type": "error",
            "answer": "dataset_id required",
        }
        
    print(f"DEBUG: [/agent/ask] dataset_id: {dataset_id}")
    try:
        from app.datasets import get_dataset
        ds = get_dataset(dataset_id)
        print(f"DEBUG: [/agent/ask] len(rows): {ds['row_count']}")
    except Exception as e:
        print(f"DEBUG: [/agent/ask] Error loading dataset length check: {e}")

    result = run_agent(req.question, dataset_id=dataset_id)
    return result


@app.post("/agent/query")
def agent_query(req: AgentAskRequest):
    """Alias for /agent/ask."""
    return agent_ask(req)


# ── Stats endpoint ──

@app.get("/api/stats")
def get_stats(dataset_id: Optional[str] = None):
    """
    Return metadata about a dataset. If dataset_id is provided,
    return stats for that dataset. Otherwise return stats for all datasets.
    """
    if dataset_id:
        try:
            ds = get_dataset(dataset_id)
            return {
                "ok": True,
                "dataset_id": dataset_id,
                "filename": ds["filename"],
                "total_rows": ds["row_count"],
                "columns": ds["columns"],
                "types": ds["types"],
                "sample_rows": ds["rows"][:6],
            }
        except KeyError as e:
            return {"ok": False, "error": str(e)}
    else:
        datasets = list_datasets()
        return {
            "ok": True,
            "datasets": datasets,
            "count": len(datasets),
        }


# ── Datasets list ──

@app.get("/api/datasets")
def api_datasets():
    """List all uploaded datasets."""
    return {"ok": True, "datasets": list_datasets()}


# ── Legacy Google Sheets endpoint (kept for backward compat) ──

@app.get("/sheet/data")
def read_sheet():
    if not _HAS_SHEETS:
        return {"ok": False, "error": "Google Sheets integration not configured."}
    try:
        rows = get_sheet_data(settings.SHEET_ID, "Products!A1:Z1000")
        return {"rows": rows}
    except Exception as e:
        return {"ok": False, "error": str(e)}